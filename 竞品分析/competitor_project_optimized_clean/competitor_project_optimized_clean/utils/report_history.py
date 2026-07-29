"""报表运行历史工具。

业务口径：
- 不再按周/小时判断是否更新；只要运行一次，就和上一轮基线做对比。
- 第一次运行只生成全量表；第二次及以后生成 上新表 / 下架表 / 全量表。
- 每个站点本轮 Excel 成功导出后，删除该站点上一轮旧报表文件，只保留最新一轮。
"""

from __future__ import annotations

import logging
import os
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def resolve_current_datetime(value: Any | None = None) -> datetime:
    """统一解析当前爬虫时间。

    支持环境变量 NEW_COLOR_TEST_NOW / SCRAPER_TEST_NOW，便于本地模拟多轮爬取。
    """
    raw = value or os.getenv("SCRAPER_TEST_NOW") or os.getenv("NEW_COLOR_TEST_NOW")
    if raw:
        text = str(raw).strip()
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                pass
        try:
            return datetime.fromisoformat(text)
        except ValueError:
            logger.warning("测试时间格式无法识别，使用当前时间: %s", text)
    return datetime.now()


def _site_report_files(output_dir: str, prefix: str) -> list[Path]:
    base = Path(output_dir)
    if not base.exists():
        return []
    return sorted(
        [p for p in base.glob(f"{prefix}*.xlsx") if p.is_file() and not p.name.startswith("~$")],
        key=lambda p: p.stat().st_mtime,
    )


def _workbook_has_diff_sheets(path: Path) -> bool:
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        names = set(wb.sheetnames)
        wb.close()
        return "上新表" in names and "下架表" in names
    except Exception as exc:
        logger.debug("读取历史报表失败，忽略该文件: %s | %s", path, exc)
        return False


def has_previous_diff_tables(output_dir: str, prefix: str) -> bool:
    """是否已经存在上一轮上新表/下架表。"""
    return any(_workbook_has_diff_sheets(path) for path in _site_report_files(output_dir, prefix))


def is_first_site_crawl(output_dir: str, prefix: str, baseline_mgr: Any | None = None) -> bool:
    """判断某站点是否第一次运行。

    主判断：是否有生命周期 baseline。baseline 为空时，没有上一轮全量可比对，只能视为第一次。
    辅助判断：如果历史报表里已经出现过「上新表/下架表」，明确不是第一次。

    说明：第一次通常只会生成全量表；第二次开始才会生成上新/下架表。
    因此这里保留 baseline 兜底，避免第二次因为还没有历史上新/下架表而被误判为第一次。
    """
    if has_previous_diff_tables(output_dir, prefix):
        return False

    if baseline_mgr is not None and hasattr(baseline_mgr, "is_empty"):
        try:
            return bool(baseline_mgr.is_empty())
        except Exception:
            pass

    return len(_site_report_files(output_dir, prefix)) == 0


def cleanup_previous_site_reports(output_dir: str, prefix: str, keep_filepath: str) -> list[str]:
    """删除同站点上一轮旧 Excel 报表，只保留本轮导出的文件。"""
    keep_path = Path(keep_filepath).resolve()
    deleted: list[str] = []

    for path in _site_report_files(output_dir, prefix):
        try:
            if path.resolve() == keep_path:
                continue
            path.unlink()
            deleted.append(str(path))
        except Exception as exc:
            logger.warning("删除旧报表失败，已跳过: %s | %s", path, exc)

    if deleted:
        logger.info("已删除 %d 个旧报表文件: %s", len(deleted), ", ".join(deleted))
    return deleted
