"""数据导出模块 - 支持多 Sheet、双层表头、冻结表头、自动列宽和按 Sheet 自定义表头。

兼容旧调用：
    export_multiple_sheets({sheet_name: records}, output_dir, header_l1=..., columns_l2=...)

新增兼容：
    export_multiple_sheets(
        {
            "主表": records,
            "颜色完整性检查": {
                "rows": audit_rows,
                "columns_l2": ["款式名", "商品名称", ...],
                "header_l1": [("颜色完整性检查", 6)],
            },
        },
        output_dir,
        header_l1=主表一级表头,
        columns_l2=主表二级表头,
    )

rows 支持：
- dataclass/对象，且有 to_row()
- dict，按 columns_l2 的字段顺序取值
- list/tuple，直接作为一行
"""

from __future__ import annotations

import logging
import os
import re
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class DataExporter:
    L1_BG_COLORS = ["4F81BD", "9BBB59", "C0504D", "8064A2", "4BACC6", "F79646"]

    @staticmethod
    def _safe_sheet_name(name: str) -> str:
        cleaned = re.sub(r"[\\/*?:\[\]]", "_", name or "Sheet")
        return cleaned[:31] or "Sheet"

    @staticmethod
    def _normalize_header_l1(
        header_l1: list[tuple[str, int]] | None,
        columns_l2: list[str],
        default_title: str = "数据",
    ) -> list[tuple[str, int]]:
        if header_l1:
            return header_l1
        return [(default_title, len(columns_l2))]

    def _write_headers(self, ws, header_l1: list[tuple[str, int]], columns_l2: list[str]) -> None:
        expected_span = sum(span for _, span in header_l1)
        if expected_span != len(columns_l2):
            raise ValueError(
                f"一级表头跨度 {expected_span} 与二级表头列数 {len(columns_l2)} 不一致"
            )

        current_col = 1
        for idx, (l1_title, span) in enumerate(header_l1):
            bg_color = self.L1_BG_COLORS[idx % len(self.L1_BG_COLORS)]
            ws.cell(row=1, column=current_col, value=l1_title)

            if span > 1:
                ws.merge_cells(
                    start_row=1,
                    start_column=current_col,
                    end_row=1,
                    end_column=current_col + span - 1,
                )

            cell = ws.cell(row=1, column=current_col)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            current_col += span

        for col_num, l2_title in enumerate(columns_l2, 1):
            cell = ws.cell(row=2, column=col_num, value=l2_title)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        ws.freeze_panes = "A3"
        ws.auto_filter.ref = f"A2:{get_column_letter(len(columns_l2))}2"

    def _adjust_column_width(self, ws) -> None:
        for col_idx, column in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            for cell in column:
                value = getattr(cell, "value", None)
                if value is not None:
                    max_length = max(max_length, len(str(value)))
            ws.column_dimensions[column_letter].width = max(10, min(max_length + 3, 60))

    @staticmethod
    def _get_value_from_dict(row: dict[str, Any], column: str) -> Any:
        """按列名取 dict 值，兼容中文列名和常见英文 key。"""
        if column in row:
            return row.get(column, "")

        aliases = {
            "网站名": "site_name",
            "品牌": "brand",
            "类目": "category",
            "排序": "current_rank",
            "排名涨跌": "rank_change",
            "商品唯一键 / SKC Key": "product_skc_key",
            "款式 ID / SPU Key": "style_spu_key",
            "款式名": "style_label",
            "商品链接": "product_url",
            "商品名称": "product_name",
            "颜色名称": "color_name",
            "尺码": "size",
            "颜色数量": "color_count",
            "颜色列表": "color_list",
            "主图": "main_image_url",
            "标价": "original_price",
            "售价": "sale_price",
            "折扣类型": "discount_type",
            "定制/现货": "stock_type",
            "商品详情描述": "detail_text",
            "爬取时间": "scrape_time",
            "数据周次": "data_week",
            "上新时间": "release_date",
            "上新类型": "new_type",
            "最近下架时间": "last_delisted_at",
            "下架前排序": "current_rank",
            "下架前标价": "original_price",
            "下架前售价": "sale_price",
            "下架前折扣类型": "discount_type",
            "最近一次出现时间": "last_seen_at",
            "下架时间": "delisted_at",
            "是否异常": "is_abnormal",
            "异常原因": "abnormal_reason",
            "代表链接": "representative_url",
        }

        alias = aliases.get(column)
        if alias and alias in row:
            return row.get(alias, "")

        return ""

    def _row_to_list(self, row: Any, columns_l2: list[str], row_index: int, sheet_name: str) -> list[Any]:
        if hasattr(row, "to_metadata") and callable(row.to_metadata):
            metadata = row.to_metadata()
            if not isinstance(metadata, dict):
                metadata = {}
            # 部分临时字段不一定写入 to_metadata，例如下架表里的 delisted_at。
            if hasattr(row, "delisted_at") and "delisted_at" not in metadata:
                metadata["delisted_at"] = getattr(row, "delisted_at", "")
            values = [self._get_value_from_dict(metadata, column) for column in columns_l2]
        elif hasattr(row, "to_row") and callable(row.to_row):
            values = list(row.to_row())
        elif isinstance(row, dict):
            values = [self._get_value_from_dict(row, column) for column in columns_l2]
        elif isinstance(row, (list, tuple)):
            values = list(row)
        else:
            values = [row]

        expected = len(columns_l2)
        if len(values) == expected:
            return values

        if len(values) < expected:
            logger.warning(
                "Sheet[%s] 第 %s 行列数不足：%s < %s，已自动补空",
                sheet_name,
                row_index,
                len(values),
                expected,
            )
            return values + [""] * (expected - len(values))

        logger.warning(
            "Sheet[%s] 第 %s 行列数过多：%s > %s，已自动截断",
            sheet_name,
            row_index,
            len(values),
            expected,
        )
        return values[:expected]

    @staticmethod
    def _resolve_sheet_payload(
        payload: Any,
        default_header_l1: list[tuple[str, int]],
        default_columns_l2: list[str],
    ) -> tuple[list[Any], list[tuple[str, int]], list[str]]:
        """支持旧 list records，也支持新 dict spec。"""
        if isinstance(payload, dict) and any(
            key in payload for key in ["rows", "records", "data", "header_l1", "columns_l2"]
        ):
            rows = payload.get("rows")
            if rows is None:
                rows = payload.get("records")
            if rows is None:
                rows = payload.get("data")
            if rows is None:
                rows = []

            columns_l2 = payload.get("columns_l2") or payload.get("columns") or default_columns_l2
            header_l1 = payload.get("header_l1") or payload.get("headers_l1")
            header_l1 = DataExporter._normalize_header_l1(header_l1, columns_l2, default_title="数据")
            return list(rows), list(header_l1), list(columns_l2)

        return list(payload or []), default_header_l1, default_columns_l2

    def export_multiple_sheets(
        self,
        sheets_data: dict[str, Any],
        output_dir: str,
        prefix: str = "report_",
        suffix: str = "",
        header_l1: list[tuple[str, int]] | None = None,
        columns_l2: list[str] | None = None,
    ) -> str:
        if header_l1 is None or columns_l2 is None:
            from utils.product_record import COLUMNS_L2, HEADER_L1_CONFIG

            columns_l2 = columns_l2 or COLUMNS_L2
            header_l1 = header_l1 or HEADER_L1_CONFIG

        default_columns_l2 = list(columns_l2)
        default_header_l1 = self._normalize_header_l1(header_l1, default_columns_l2)

        os.makedirs(output_dir, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{prefix}{timestamp}{suffix}.xlsx"
        filepath = os.path.join(output_dir, filename)

        # 如果本次运行存在最终失败的 retry tasks，自动追加一个排查 sheet。
        # 不影响原有业务 sheet；只有存在 failed tasks 时才会出现。
        if os.getenv("RETRY_DEAD_LETTER_ENABLED", "true").strip().lower() in {"1", "true", "yes", "y"}:
            try:
                from utils.retry_queue import collect_failed_retry_rows

                retry_failed_rows = collect_failed_retry_rows()
            except Exception:
                retry_failed_rows = []

            if retry_failed_rows and "Retry_Failed_Tasks" not in sheets_data:
                sheets_data = dict(sheets_data)
                sheets_data["Retry_Failed_Tasks"] = {
                    "rows": retry_failed_rows,
                    "columns_l2": [
                        "run_id",
                        "site_key",
                        "task_type",
                        "identity_key",
                        "attempts",
                        "max_attempts",
                        "status",
                        "last_error",
                        "created_at",
                        "updated_at",
                        "failed_file",
                    ],
                    "header_l1": [("Retry Failed Tasks", 11)],
                }

        wb = Workbook()
        wb.remove(wb.active)

        used_sheet_names: set[str] = set()

        for raw_sheet_name, payload in sheets_data.items():
            sheet_name = self._safe_sheet_name(raw_sheet_name)
            base_sheet_name = sheet_name
            suffix_idx = 2
            while sheet_name in used_sheet_names:
                tail = f"_{suffix_idx}"
                sheet_name = f"{base_sheet_name[:31 - len(tail)]}{tail}"
                suffix_idx += 1
            used_sheet_names.add(sheet_name)

            rows, sheet_header_l1, sheet_columns_l2 = self._resolve_sheet_payload(
                payload,
                default_header_l1,
                default_columns_l2,
            )

            ws = wb.create_sheet(title=sheet_name)
            self._write_headers(ws, sheet_header_l1, sheet_columns_l2)

            for row_index, record in enumerate(rows, start=1):
                ws.append(self._row_to_list(record, sheet_columns_l2, row_index, sheet_name))

            self._adjust_column_width(ws)

        wb.save(filepath)
        logger.info("🎉 导出完成：%s", filepath)
        return filepath
